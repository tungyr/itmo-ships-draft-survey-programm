import sys
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog, QMessageBox
from PyQt5.QtGui import QIcon
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import date, datetime, time

labels = ["Fwd mean draft, m: ",
          "Middle mean draft, m: ",
          "Aft mean draft, m: ",
          "Fwd mark misplacement, m: ",
          "Mid mark misplacement, m: ",
          "Aft mark misplacement, m: ",
          "Apparent trim, m: ",
          "Fwd draft correction, m: ",
          "Mid draft correction, m: ",
          "Aft draft correction, m: ",
          "Fwd corrected draft, m: ",
          "Mid corrected draft, m: ",
          "Aft corrected draft, m: ",
          "True trim, m: ",
          "Deflection: ",
          "Mean of means corrected, m:",
          "Displacement by MOMC, mt: ",
          "TPC, mt: ",
          "LCF, m: ",
          "First trim correction, mt:",
          "MTC by MOMC: ",
          "MTC +: ",
          "MTC -: ",
          "MTC difference: ",
          "Second trim correction, mt:",
          "Disp. corrected by trim, mt: ",
          "Constant, mt: ",
          "Displacement corrected, mt: "]


class App(QWidget):

    def __init__(self, data, vessel_name):
        super().__init__()
        self.title = 'PyQt5 file save dialog'
        self.left = 100
        self.top = 100
        self.width = 640
        self.height = 480
        self.data = data
        self.filename = ''
        self.vessel_name = vessel_name
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        # self.openFileNameDialog()
        # self.openFileNamesDialog()
        self.saveFileDialog()

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        time_stamp = str(datetime.today())
        time_stamp = f'{time_stamp[:13]}-{time_stamp[14:16]}'
        self.filename, _ = QFileDialog.getSaveFileName(self, "Save File", f"{time_stamp} - "
                                                        f"{self.vessel_name} draft calculation" + ".xlsx",
                                                  "Excel Files (*.xlsx);;All Files (*)", options=options)

        if self.filename != "":
            if '.xlsx' not in self.filename:
                self.filename = self.filename + ".xlsx"
                print(self.filename)
            self.save_xls()

        return self.filename

    def save_xls(self):

        alignment_center = Alignment(horizontal='center')

        wb = Workbook()
        ws = wb.active

        ws.merge_cells('A1:I1')
        header = ws["A1"]
        header.value = f'Draft-survey calculation of m/v "{self.vessel_name}"'
        header.font = Font(name='Calibri', size=16, bold=True)
        header.alignment = alignment_center

        ws["E2"] = str(datetime.today())[:16]

        ws.merge_cells('A4:I4')
        results_title = ws["A4"]
        results_title.value = 'Results: '
        results_title.alignment = alignment_center
        results_title.font = Font(name="Calibri", size=14, bold=True)

        row = 6
        for label in labels:
            cell = ws["B" + str(row)]
            cell.value = label
            cell.font = Font(name="Calibri", size=14)
            row += 1

        row = 6
        for item in self.data:
            cell = ws["G" + str(row)]
            cell.value = str(item)
            cell.font = Font(name="Calibri", size=14)
            row += 1

        try:
            wb.save(filename=self.filename)
        except PermissionError:
            warn = QMessageBox.warning(self, 'Error!',
                                       "Error! File to overwrite is open!"
                                       + "\n", QMessageBox.Ok)

        self.close()




if __name__ == "__main__":

    app = QApplication(sys.argv)
    ex = App()
    ex.show()
    ex.close()



